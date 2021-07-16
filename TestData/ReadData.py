import openpyxl

excel_file = openpyxl.load_workbook('C://Users/rebecca.madeira/Desktop/RobotAutomationTest/ExcelFiles/TestingData.xlsx')

def number_rows(sheetname):
    sh = excel_file[sheetname]
    return sh.max_row

def fetch_cell_data(sheetname, row, col):
    sh = excel_file[sheetname]
    cell = sh.cell(int(row),int(col))
    return cell.value

